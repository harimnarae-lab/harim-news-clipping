"""
하림산업 뉴스 클리핑 — Streamlit 웹앱 버전
실행: streamlit run app.py
"""

import streamlit as st
import requests
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import re
import io

# ══════════════════════════════════════════════════════════════
# API 설정
# ══════════════════════════════════════════════════════════════
CLIENT_ID     = "tFPDiFEQUVsPxGyAVFIW"
CLIENT_SECRET = "idIRmkiThO"

# ══════════════════════════════════════════════════════════════
# 키워드 정의 (기존 스크립트 그대로)
# ══════════════════════════════════════════════════════════════
SECTION_자사 = {
    '하림산업': '하림산업',
    '더미식': '더미식', 'the미식': '더미식',
    '더미식 라면': '더미식', '더미식 즉석밥': '더미식',
    '더미식 만두': '더미식', '더미식 국물': '더미식',
    '더미식 밀키트': '더미식', '더미식 볶음밥': '더미식',
    '푸디버디': '푸디버디', '푸디버디 라면': '푸디버디',
    '푸디버디 핫도그': '푸디버디', '푸디버디 튀김': '푸디버디',
    '멜팅피스': '맥시칸', '맥시칸 치킨': '맥시칸',
    '퍼스트키친': '퍼스트키친',
    '하림산업 HMR': 'HMR', '하림산업 간편식': '간편식',
}

SECTION_그룹 = {
    '하림그룹': '하림그룹', '하림지주': '하림지주',
    '김홍국': '회장', 'NS홈쇼핑 푸드': 'NS홈쇼핑',
    'NS푸드페스타': 'NS홈쇼핑', '오드그로서': '오드그로서',
    '팬오션 식품': '팬오션', '하림 공장 투어': '공장투어',
    '하림 익산': '익산/행사', '하림 팝업스토어': '팝업스토어',
    '용가리 팝업': '팝업스토어', '하림 협업': '협업',
}

SECTION_경쟁사 = {
    '농심 라면': '라면', '삼양 라면': '라면', '오뚜기 라면': '라면',
    '팔도 라면': '라면', '농심 신라면': '라면', '불닭볶음면': '라면',
    '농심 글로벌': '글로벌', '삼양식품 글로벌': '글로벌',
    '오뚜기 해외': '글로벌', 'CJ제일제당 해외': '글로벌',
    'K푸드 수출': '글로벌', '비비고 해외': '글로벌',
    '비비고 만두': '만두', '풀무원 만두': '만두',
    '동원 냉동식품': '냉동식품', '신세계푸드 HMR': 'HMR',
    'CJ 햇반': '즉석밥', '오뚜기 즉석밥': '즉석밥',
    '가정간편식 시장': 'HMR', 'HMR 시장': 'HMR', '밀키트 시장': '밀키트',
    '청정원 소스': '소스/양념', '삼양 소스': '소스/양념',
    '불닭소스': '소스/양념', '덮밥소스': '소스/양념',
    '농심 인사': '인사', 'CJ제일제당 인사': '인사', '오뚜기 인사': '인사',
    '농심 경영': '기업동향', 'CJ제일제당 경영': '기업동향',
    '삼양식품 경영': '기업동향', '풀무원 경영': '기업동향',
    '동원 경영': '기업동향', '신세계푸드 경영': '기업동향',
    '식품업체 수상': '수상', '청정원 수상': '수상',
}

SECTION_유통 = {
    '식품업계': '식품업계', '식품업계 원가': '식품업계',
    '식품업계 물가': '물가동향', '라면값': '물가동향',
    '식품 물가': '물가정책', '식품 가격 인하': '물가정책',
    'GMO 표시': '식품업계/규제이슈', 'GMO 완전표시': '식품업계/규제이슈',
    'NON-GMO': '식품업계/규제이슈', '식품 담합': '식품업계/규제이슈',
    '식품 규제': '식품업계/규제이슈',
    '편의점 간편식': '편의점', '편의점 HMR': '편의점',
    '편의점 트렌드': '편의점', 'GS25 신제품': '편의점',
    'CU 신제품': '편의점', '세븐일레븐 신제품': '편의점',
    '이마트 식품': '대형마트', '롯데마트 식품': '대형마트', '홈플러스 식품': '대형마트',
    '쿠팡 식품': '이커머스', '컬리 식품': '이커머스', 'SSG 식품': '이커머스',
    '온라인 식품': '온라인몰', '새벽배송 식품': '이커머스',
    '외식업계': '외식업계', '급식업계': '급식업계', '단체급식': '급식업계',
}

SECTION_부정 = {
    '하림산업 적자': '경영성과', '하림산업 논란': '논란',
    '하림산업 위생': '위생/안전', '더미식 논란': '논란',
    '더미식 위생': '위생/안전', '더미식 리콜': '리콜/안전',
    '푸디버디 논란': '논란', '맥시칸 논란': '논란',
    '하림산업 갑질': '갑질/노무', '하림산업 불매': '불매운동',
    '하림 재무': '경영성과', '하림산업 부실': '경영성과',
}

ALL_KEYWORDS = {}
for kw, cat in SECTION_자사.items():    ALL_KEYWORDS[kw] = ('[자사]', cat)
for kw, cat in SECTION_그룹.items():   ALL_KEYWORDS[kw] = ('[그룹 및 계열사]', cat)
for kw, cat in SECTION_경쟁사.items(): ALL_KEYWORDS[kw] = ('[경쟁사]', cat)
for kw, cat in SECTION_유통.items():   ALL_KEYWORDS[kw] = ('[식품 및 유통업계]', cat)
for kw, cat in SECTION_부정.items():   ALL_KEYWORDS[kw] = ('[자사 부정기사]', cat)

PRESS_MAPPING = {
    'chosun.com': '조선일보', 'biz.chosun.com': '조선비즈',
    'it.chosun.com': 'IT조선', 'joongang.co.kr': '중앙일보',
    'news.joins.com': '중앙일보', 'donga.com': '동아일보',
    'hani.co.kr': '한겨레', 'khan.co.kr': '경향신문',
    'hankyung.com': '한국경제', 'mk.co.kr': '매일경제',
    'sedaily.com': '서울경제', 'mt.co.kr': '머니투데이',
    'edaily.co.kr': '이데일리', 'asiae.co.kr': '아시아경제',
    'newsis.com': '뉴시스', 'news1.kr': '뉴스1',
    'yonhapnews.co.kr': '연합뉴스', 'yna.co.kr': '연합뉴스',
    'heraldcorp.com': '헤럴드경제', 'biz.heraldcorp.com': '헤럴드경제',
    'fnnews.com': '파이낸셜뉴스', 'newspim.com': '뉴스핌',
    'ajunews.com': '아주경제', 'businesspost.co.kr': '비즈니스포스트',
    'newsway.co.kr': '뉴스웨이', 'dt.co.kr': '디지털타임스',
    'etnews.com': '전자신문', 'sbs.co.kr': 'SBS',
    'kbs.co.kr': 'KBS', 'mbc.co.kr': 'MBC',
    'jtbc.co.kr': 'JTBC', 'ytn.co.kr': 'YTN',
    'naver.com': '네이버', 'daum.net': '다음', 'v.daum.net': '다음',
    'sisajournal-e.com': '시사저널e', 'sisajournal.com': '시사저널',
    'dailian.co.kr': '데일리안', 'daily.hankooki.com': '데일리한국',
    'theguru.co.kr': '더구루', 'newdaily.co.kr': '뉴데일리',
    'ekn.kr': '에너지경제', 'foodnews.co.kr': '식품음료신문',
    'foodbank.co.kr': '식품외식경제', 'inews24.com': '아이뉴스24',
    'tf.co.kr': '더팩트', 'asiatoday.co.kr': '아시아투데이',
    'seoul.co.kr': '서울신문', 'dnews.co.kr': '대한경제',
    'ddaily.co.kr': '디지털데일리', 'hansbiz.co.kr': '한스경제',
    'econovill.com': '이코노믹리뷰', 'g-enews.com': '글로벌이코노믹',
    'thebigdata.co.kr': '빅데이터뉴스', 'youthdaily.co.kr': '청년일보',
    'fetv.co.kr': 'FETV', 'metroseoul.co.kr': '메트로',
    'nocutnews.co.kr': '노컷뉴스', 'kukinews.com': '쿠키뉴스',
    'consumernews.co.kr': '소비자가만드는신문',
    'kookje.co.kr': '국제신문', 'imaeil.com': '매일신문',
    'jjan.kr': '전북일보', 'domin.co.kr': '전북도민일보',
    'agrinet.co.kr': '농업인신문', 'agrihanguk.com': '한국농업신문',
    'newsfarm.co.kr': '농업경제신문',
}

# ══════════════════════════════════════════════════════════════
# 핵심 함수 (기존 로직 그대로)
# ══════════════════════════════════════════════════════════════
def get_press_name(url):
    try:
        domain = url.split('/')[2].replace('www.', '')
        for key, name in PRESS_MAPPING.items():
            if key in domain:
                return name
        return domain
    except:
        return '알 수 없음'

def convert_url(url):
    try:
        if '//m.' in url:      url = url.replace('//m.', '//www.')
        if '//mobile.' in url: url = url.replace('//mobile.', '//www.')
        for mob, web in {
            'n.news.naver.com': 'news.naver.com',
            'm.khan.co.kr': 'www.khan.co.kr',
            'm.mk.co.kr': 'www.mk.co.kr',
            'm.mt.co.kr': 'www.mt.co.kr',
        }.items():
            if mob in url: url = url.replace(mob, web)
        return url
    except:
        return url

def clean(text):
    text = re.sub(r'<[^>]+>', '', text)
    return (text.replace('&quot;', '"').replace('&apos;', "'")
                .replace('&amp;', '&').replace('&lt;', '<')
                .replace('&gt;', '>').replace('\xa0', ' ').strip())

def search_keyword(keyword, start_date, end_date):
    section, category = ALL_KEYWORDS.get(keyword, ('[경쟁사]', keyword))
    headers = {
        'X-Naver-Client-Id': CLIENT_ID,
        'X-Naver-Client-Secret': CLIENT_SECRET,
    }
    params = {'query': keyword, 'display': 100, 'start': 1, 'sort': 'date'}
    articles = []
    try:
        resp = requests.get(
            "https://openapi.naver.com/v1/search/news.json",
            headers=headers, params=params, timeout=10
        )
        resp.raise_for_status()
        for item in resp.json().get('items', []):
            try:
                pub = datetime.strptime(
                    item.get('pubDate', ''), '%a, %d %b %Y %H:%M:%S %z'
                ).replace(tzinfo=None)
                if not (start_date <= pub <= end_date):
                    continue
                orig = item.get('originallink', '')
                articles.append({
                    'section': section,
                    'category': category,
                    '헤드라인': clean(item.get('title', '')),
                    '매체': get_press_name(orig) if orig else '알 수 없음',
                    '보도일': pub,
                    'URL': convert_url(item.get('link', '')),
                })
            except:
                continue
    except Exception as e:
        pass
    return articles

def create_excel_bytes(all_articles, end_date):
    """엑셀 파일을 bytes로 반환 (다운로드용)"""
    unique, seen = [], set()
    for a in all_articles:
        if a['URL'] not in seen:
            seen.add(a['URL'])
            unique.append(a)

    wb = Workbook()
    ws = wb.active
    now = datetime.now()
    ws.title = f"{now.month}월{now.day}일"

    def fill_cell(hex_):
        return PatternFill(start_color=hex_, end_color=hex_, fill_type='solid')

    def border():
        t = Side(style='thin', color='D0D0D0')
        return Border(left=t, right=t, top=t, bottom=t)

    HEADER_BG  = 'F2F2F2'
    LINK_COLOR = '0000FF'
    NEG_BG     = 'FFFF00'
    BASE_FONT  = Font(name='맑은 고딕', size=9)
    BOLD_FONT  = Font(name='맑은 고딕', size=9, bold=True)

    def w(cell, val, font=None, bg=None, align=None, b=True, hyper=None):
        cell.value = val
        cell.font = font or BASE_FONT
        if bg: cell.fill = fill_cell(bg)
        cell.alignment = align or Alignment(horizontal='left', vertical='center')
        if b: cell.border = border()
        if hyper: cell.hyperlink = hyper

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 13
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 24

    row = 1
    MAIN_SECTIONS = ['[자사]', '[그룹 및 계열사]', '[경쟁사]', '[식품 및 유통업계]']

    for section_label in MAIN_SECTIONS:
        sec_arts = [a for a in unique if a['section'] == section_label]
        ws.row_dimensions[row].height = 6
        ws[f'A{row}'].value = '-'
        ws[f'A{row}'].font = BASE_FONT
        row += 1

        ws.row_dimensions[row].height = 16
        c = ws[f'A{row}']
        c.value = section_label
        c.font = BOLD_FONT
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

        ws.row_dimensions[row].height = 16
        for col, txt in zip('ABCDEF', ['NO.', '카테고리', '보도일', ' 헤드라인 ', '매체', ' 비고']):
            w(ws[f'{col}{row}'], txt, font=BOLD_FONT, bg=HEADER_BG,
              align=Alignment(horizontal='center', vertical='center'))
        row += 1

        if not sec_arts:
            ws.row_dimensions[row].height = 14
            for col in 'ABCDEF':
                w(ws[f'{col}{row}'], '-',
                  align=Alignment(horizontal='center', vertical='center'))
            row += 1
            continue

        for i, a in enumerate(sec_arts, 1):
            ws.row_dimensions[row].height = 16
            center = Alignment(horizontal='center', vertical='center')
            left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
            w(ws[f'A{row}'], i, align=center)
            w(ws[f'B{row}'], a['category'], align=center)
            dc = ws[f'C{row}']
            dc.value = a['보도일']
            dc.number_format = 'YYYY.MM.DD'
            dc.font = BASE_FONT
            dc.alignment = center
            dc.border = border()
            w(ws[f'D{row}'], a['헤드라인'],
              font=Font(name='맑은 고딕', size=9, color=LINK_COLOR),
              align=left, hyper=a['URL'])
            w(ws[f'E{row}'], a['매체'], align=center)
            w(ws[f'F{row}'], '', align=left)
            row += 1

    # 부정기사
    neg_arts = [a for a in unique if a['section'] == '[자사 부정기사]']
    ws.row_dimensions[row].height = 8
    row += 1

    neg_title = ws[f'A{row}']
    neg_title.value = '부정기사 모니터링'
    neg_title.font  = Font(name='맑은 고딕', size=11, bold=True)
    neg_title.fill  = fill_cell(NEG_BG)
    neg_title.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(f'A{row}:F{row}')
    ws.row_dimensions[row].height = 18
    row += 1

    c = ws[f'A{row}']
    c.value = '[자사 부정기사]'
    c.font = BOLD_FONT
    ws.merge_cells(f'A{row}:F{row}')
    row += 1

    ws.row_dimensions[row].height = 16
    for col, txt in zip('ABCDEF', ['NO.', '카테고리', '보도일', ' 헤드라인 ', '매체', ' 비고']):
        w(ws[f'{col}{row}'], txt, font=BOLD_FONT, bg=HEADER_BG,
          align=Alignment(horizontal='center', vertical='center'))
    row += 1

    if not neg_arts:
        for col in 'ABCDEF':
            w(ws[f'{col}{row}'], '-',
              align=Alignment(horizontal='center', vertical='center'))
        row += 1
    else:
        for i, a in enumerate(neg_arts, 1):
            center = Alignment(horizontal='center', vertical='center')
            left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
            w(ws[f'A{row}'], i, align=center)
            w(ws[f'B{row}'], a['category'], align=center)
            dc = ws[f'C{row}']
            dc.value = a['보도일']
            dc.number_format = 'YYYY.MM.DD'
            dc.font = BASE_FONT
            dc.alignment = center
            dc.border = border()
            w(ws[f'D{row}'], a['헤드라인'],
              font=Font(name='맑은 고딕', size=9, color='FF0000'),
              align=left, hyper=a['URL'])
            w(ws[f'E{row}'], a['매체'], align=center)
            w(ws[f'F{row}'], '', align=left)
            row += 1

    ws.freeze_panes = 'A2'

    # 파일을 메모리에 저장 후 bytes 반환
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), len(unique)

# ══════════════════════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="하림산업 뉴스 클리핑",
    page_icon="🍜",
    layout="centered"
)

st.title("🍜 하림산업 뉴스 클리핑")
st.caption("네이버 뉴스 API 기반 자동 수집 · 엑셀 다운로드")

st.divider()

# ── 검색 기간 설정 ──────────────────────────────────────────
st.subheader("📅 검색 기간 설정")

now = datetime.now()
w = now.weekday()
if   w == 5: auto_days = 1
elif w == 6: auto_days = 2
elif w == 0: auto_days = 3
else:        auto_days = 1

auto_start = (now - timedelta(days=auto_days)).replace(hour=9, minute=0, second=0, microsecond=0)

col1, col2 = st.columns(2)
with col1:
    start_date_input = st.date_input("시작 날짜", value=auto_start.date())
    start_time_input = st.time_input("시작 시간", value=auto_start.time())
with col2:
    end_date_input = st.date_input("종료 날짜", value=now.date())
    end_time_input = st.time_input("종료 시간", value=now.time())

start_dt = datetime.combine(start_date_input, start_time_input)
end_dt   = datetime.combine(end_date_input,   end_time_input)

st.info(f"검색 기간: **{start_dt.strftime('%Y.%m.%d %H:%M')}** ~ **{end_dt.strftime('%Y.%m.%d %H:%M')}**")

st.divider()

# ── 키워드 구성 확인 ─────────────────────────────────────────
with st.expander("🔍 검색 키워드 구성 보기"):
    for sec in ['[자사]', '[그룹 및 계열사]', '[경쟁사]', '[식품 및 유통업계]', '[자사 부정기사]']:
        kws = [k for k, (s, _) in ALL_KEYWORDS.items() if s == sec]
        st.write(f"**{sec}**: {len(kws)}개")
    st.write(f"→ **총 {len(ALL_KEYWORDS)}개** 키워드")

st.divider()

# ── 실행 버튼 ────────────────────────────────────────────────
if st.button("🚀 뉴스 수집 시작", type="primary", use_container_width=True):

    if start_dt >= end_dt:
        st.error("종료 시간이 시작 시간보다 앞에 있습니다. 다시 확인해주세요.")
    else:
        all_articles = []
        sections_order = ['[자사]', '[그룹 및 계열사]', '[경쟁사]', '[식품 및 유통업계]', '[자사 부정기사]']

        # 진행 상황 표시
        progress_bar = st.progress(0)
        status_text  = st.empty()
        total_kw     = len(ALL_KEYWORDS)
        done         = 0

        for sec in sections_order:
            keywords_in_sec = [k for k, (s, _) in ALL_KEYWORDS.items() if s == sec]
            for kw in keywords_in_sec:
                status_text.text(f"검색 중: {sec} › {kw}")
                arts = search_keyword(kw, start_dt, end_dt)
                all_articles.extend(arts)
                done += 1
                progress_bar.progress(done / total_kw)
                time.sleep(0.1)

        progress_bar.progress(1.0)
        status_text.text("✅ 수집 완료! 엑셀 파일 생성 중...")

        # 엑셀 생성
        excel_bytes, unique_count = create_excel_bytes(all_articles, end_dt)

        status_text.empty()
        progress_bar.empty()

        # 결과 요약
        st.success(f"✅ 수집 완료! 총 **{unique_count}건** (중복 제거 후)")

        # 섹션별 건수
        for sec in sections_order:
            cnt = len([a for a in all_articles if a['section'] == sec])
            st.write(f"  · {sec}: {cnt}건")

        st.divider()

        # 다운로드 버튼
        filename = f"{now.strftime('%Y%m%d_%H%M')}_하림산업_뉴스클리핑.xlsx"
        st.download_button(
            label="⬇️ 엑셀 다운로드",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
